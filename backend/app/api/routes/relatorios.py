from fastapi import APIRouter, Depends, Query, Response
from datetime import date
from typing import Optional
import io
from ...core.auth import get_current_user
from ...core.database import get_service_client
from ...schemas.dashboard import FiltrosDashboard
from ...services import kpi_service

router = APIRouter(prefix="/relatorios", tags=["relatorios"])


def _build_filtros(
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    vendedor: Optional[str] = Query(None),
    gestor: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
) -> FiltrosDashboard:
    return FiltrosDashboard(
        data_inicio=data_inicio, data_fim=data_fim,
        vendedor=vendedor, gestor=gestor, categoria=categoria,
    )


@router.get("/export/csv")
def export_csv(
    tipo: str = Query("vendas"),
    filtros: FiltrosDashboard = Depends(_build_filtros),
    _user: dict = Depends(get_current_user),
):
    import csv
    db = get_service_client()

    if tipo == "dde":
        rows = db.table("dde_itens").select("*").order("data_referencia", desc=True).limit(500).execute().data or []
    else:
        _, df, _ = kpi_service._get_enriched(filtros)
        if df.empty:
            return Response(content="Sem dados.", media_type="text/plain")

        export_cols = {
            "vendas":    ["data_referencia", "cliente_nome", "descricao", "categoria", "familia",
                          "vendedor_nome", "supervisor", "rota", "rede",
                          "quantidade", "preco_cx", "venda_bruta", "venda_liquida",
                          "custo_unitario", "margem_bruta", "margem_pct", "desconto"],
            "ruptura":   ["descricao", "categoria", "quantidade", "estoque_atual", "ruptura_prevista"],
            "logistica": ["rota", "vendedor_nome", "cliente_nome", "venda_liquida", "quantidade"],
            "margem":    ["descricao", "categoria", "vendedor_nome", "venda_liquida", "margem_bruta", "margem_pct"],
        }
        cols = [c for c in export_cols.get(tipo, export_cols["vendas"]) if c in df.columns]

        if tipo == "ruptura":
            df = df[df["ruptura_prevista"] < 0].copy()
            df["ruptura_prevista"] = df["ruptura_prevista"].abs()

        rows = df[cols].round(4).to_dict(orient="records")

    if not rows:
        return Response(content="Sem dados.", media_type="text/plain")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)

    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo}.csv"},
    )


@router.get("/export/excel")
def export_excel(
    tipo: str = Query("vendas"),
    filtros: FiltrosDashboard = Depends(_build_filtros),
    _user: dict = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    db = get_service_client()

    if tipo == "dde":
        rows = db.table("dde_itens").select("*").order("data_referencia", desc=True).limit(500).execute().data or []
    else:
        _, df, _ = kpi_service._get_enriched(filtros)
        if df.empty:
            rows = []
        else:
            export_cols = {
                "vendas":    ["data_referencia", "cliente_nome", "descricao", "categoria",
                              "vendedor_nome", "supervisor", "rota", "rede",
                              "quantidade", "preco_cx", "venda_bruta", "venda_liquida",
                              "margem_bruta", "margem_pct", "desconto"],
                "ruptura":   ["descricao", "categoria", "quantidade", "estoque_atual", "ruptura_prevista"],
                "logistica": ["rota", "vendedor_nome", "cliente_nome", "venda_liquida", "quantidade"],
                "margem":    ["descricao", "categoria", "vendedor_nome", "venda_liquida", "margem_bruta", "margem_pct"],
            }
            cols = [c for c in export_cols.get(tipo, export_cols["vendas"]) if c in df.columns]
            if tipo == "ruptura":
                df = df[df["ruptura_prevista"] < 0].copy()
                df["ruptura_prevista"] = df["ruptura_prevista"].abs()
            rows = df[cols].round(4).to_dict(orient="records")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    if rows:
        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="004400")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(h) for h in headers])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo}.xlsx"},
    )


@router.get("/diario")
def relatorio_diario(
    data_referencia: Optional[date] = Query(None),
    _user: dict = Depends(get_current_user),
):
    ref = data_referencia or date.today()
    filtros = FiltrosDashboard(data_inicio=ref, data_fim=ref)
    _, df, _ = kpi_service._get_enriched(filtros)

    if df.empty:
        return {
            "data_referencia": str(ref),
            "faturamento_liquido": 0,
            "margem_bruta_rs": 0,
            "margem_bruta_pct": 0,
            "ruptura_kg": 0,
            "qtd_itens": 0,
            "por_vendedor": [],
        }

    fat_liq = float(df["venda_liquida"].sum())
    margem = float(df["margem_bruta"].sum())
    ruptura = float(df[df["ruptura_prevista"] < 0]["ruptura_prevista"].abs().sum())

    por_vendedor_g = (
        df.groupby("vendedor_nome")
        .agg(venda=("venda_liquida", "sum"), margem=("margem_bruta", "sum"))
        .reset_index()
        .sort_values("venda", ascending=False)
    )

    return {
        "data_referencia": str(ref),
        "faturamento_liquido": round(fat_liq, 2),
        "margem_bruta_rs": round(margem, 2),
        "margem_bruta_pct": round(margem / fat_liq * 100, 2) if fat_liq > 0 else 0,
        "ruptura_kg": round(ruptura, 3),
        "qtd_itens": len(df),
        "por_vendedor": [
            {
                "vendedor": row["vendedor_nome"],
                "venda": round(float(row["venda"]), 2),
                "margem": round(float(row["margem"]), 2),
            }
            for _, row in por_vendedor_g.iterrows()
        ],
    }
